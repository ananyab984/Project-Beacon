"""
Stage 4: join ada_projectbeacon_output.json (one record per unique
Profile_Link) back onto every Case_ID row in ADA_Public_Profile_Dataset.xlsx,
then write the result into a copy of Template_ProjectBeacon.xlsx, preserving
its header rows/formatting and appending one data row per case starting at
row 5.
"""

import json

import openpyxl
import pandas as pd

INPUT_DATASET = "ADA_Public_Profile_Dataset.xlsx"
PARSED_JSON = "ada_projectbeacon_output.json"
TEMPLATE_PATH = "../Template_ProjectBeacon.xlsx"
OUTPUT_PATH = "ADA_ProjectBeacon_Output.xlsx"

# Column order matches Template_ProjectBeacon.xlsx row 3, columns B-P.
TEMPLATE_FIELD_ORDER = [
    "Reachout Date",
    "Application Date",
    "First_Name",
    "Full_Name",
    "Country_of_Residence",
    "Source",
    "Profile_Link",
    "Contact_Number",
    "Email_Address",
    "Services",
    "Source_Language",
    "Target_Language",
    "Secondary_Languages",
    "Years_of_Exp",
    "Vendor_Experience",
]


def main():
    cases = pd.read_excel(INPUT_DATASET)

    with open(PARSED_JSON, "r", encoding="utf-8") as f:
        parsed_records = json.load(f)
    by_link = {r["Profile_Link"]: r for r in parsed_records}

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb["Sheet1"]

    # Column A is the template's row-label column ("Field>>"/"Example>>"),
    # not a data field, so it's left blank on data rows. Case_ID is appended
    # in column Q, one past the template's 15 official fields (B-P), for
    # traceability back to the source case list without altering the
    # template's own column layout.
    ws.cell(row=3, column=17, value="Case_ID")

    row_idx = 5
    matched, unmatched = 0, 0
    for _, case in cases.iterrows():
        link = case["Profile_Link"]
        record = by_link.get(link)
        if record is None:
            unmatched += 1
            print(f"WARNING: no parsed record for {case['Case_ID']} ({link})")
            record = {field: None for field in TEMPLATE_FIELD_ORDER}
            record["Source"] = "ADA"
            record["Profile_Link"] = link
        else:
            matched += 1

        for col_offset, field in enumerate(TEMPLATE_FIELD_ORDER, start=2):
            ws.cell(row=row_idx, column=col_offset, value=record.get(field))
        ws.cell(row=row_idx, column=17, value=case["Case_ID"])
        row_idx += 1

    wb.save(OUTPUT_PATH)
    print(f"\nMatched {matched} cases, {unmatched} unmatched.")
    print(f"Saved {row_idx - 5} rows to {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
